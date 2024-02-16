'''file formate Excel Workbook'''
from openpyxl import Workbook, load_workbook

'''some what automating the process'''

source_inp=input('Enter the path of the source file:')
source_inp=source_inp[1:len(source_inp)-1]
destination=input ('Enter the name you want the output file to be: ')+'.xlsx'
path=""
inp_name=""
count=0
for i in range(len(source_inp)-1,-1,-1):
    if source_inp[i]!='\\' and count==0:
        inp_name=source_inp[i]+inp_name
    else:
        count=1
    if count==1:
        if source_inp[i]=='\\':
            path='/'+path
        else:
            path=source_inp[i]+path

a=path+destination
print(source_inp)
wb = Workbook()
wb.save(filename=destination)
''''''''''''''''''''''''''''''''''''''''''


# loading the source xl
actual = load_workbook(source_inp)
# take_from = actual.active


# loading the output xl
"""needs to be created a blankfile with the exact name and formate beforhand """
to = load_workbook(destination)
write_to = to.active


# dic_store for future sorting by district name
# writecount for writing the final result
dic_store={}
writecount = 1

sheetnames=list(actual.sheetnames)

for activesheet in sheetnames:
    # variable empty for counting empty cells and count for taking inpur row-wise
    empty = 0
    count = 1   
    take_from=actual[activesheet]
    while empty<100:
        name=take_from[f'a{count}'].value
        memberid=take_from[f'b{count}'].value
        # other=take_from[f'c{count}'].value
        mobile=take_from[f'g{count}'].value
        dist=take_from[f'j{count}'].value
        count+=1
        
        if memberid==None and name ==None:  #and other==None 
            empty+=1
        else:
            mobile=str(mobile)
            # removing unwanted items from the name
            try:
                while True:

                    if "(" in name:

                        start=name.index('(')
                        end=name.index(')')
                        new=''
                        for i in range(len(name)):
                            if i <start or  i>end:
                                new+=name[i]
                        name=new
                    elif 'LIFE MEMBER' in name:
                        name=name.replace('LIFE MEMBER','')
                    else:
                        break  
            except:
                pass
            
            # adding prefix and suffix to the name
            try:
                name=name.strip()
                name="Ld. "+ name +' Sir, '
            except:
                pass
            
            # sorting the extra numbers according to different parameters
            if',' in mobile:
                mobile=mobile.split(',')
            elif '/' in mobile:
                mobile=mobile.split('/')
            elif '-' in mobile :
                mobile=mobile.split('-')
            elif len(mobile)==22:
                mobile=[mobile[0:11],mobile[11:22]]
            elif len(mobile)==21:
                mobile=[f'0{mobile[0:10]}',mobile[10:21]]
            else:
                mobile=mobile.split()
            # stripper
            for i in range(len(mobile)):
                mobile[i]=mobile[i].strip()
            
            # fixing the mobile list.... cause inputs are fucked and sepereted by various sorts of shits
            for i in range(len(mobile)):
                if len(mobile[i]) == 10 and mobile[i][0] == '1':
                    mobile[i] = '0' + mobile[i]
                if len(mobile[i]) == 12 and mobile[i][-1] == '.':
                    mobile[i] = mobile[i][0:11]
                if len(mobile[i]) == 12 and mobile[i][-1] == ',':
                    mobile[i] = mobile[i][0:11]
                if len(mobile[i]) == 12 and mobile[i][0] == '`':
                    mobile[i] = mobile[i][1:12]
            
            # replacing any invalid numbers of citycell numbers with '0'
            for i in mobile:
                if (len(i)!= 11) or ('011' == i[0:3]):
                    print(activesheet, name, memberid,mobile[mobile.index(i)])
                    mobile[mobile.index(i)] = 0
            
            # creating new mobile without '0's  --- in other words without any invalid numbers
            mobile = [i for i in mobile if i != 0]

            # finally writing the valid numbers in the desired file
            if mobile != []:
                for i in mobile:
                    write_to[f'a{writecount}'] = memberid
                    write_to[f'b{writecount}'] = name
                    # write_to[f'c{writecount}'] = other
                    write_to[f'd{writecount}'] = i
                    write_to[f'e{writecount}'] = dist
                    writecount += 1

            # print(activesheet,count)

to.save(destination)
