'''file formate Excel Workbook'''
from openpyxl import Workbook, load_workbook

'''some what automating the process'''
source_inp=input('Enter the path of the source file:')
destination=input ('Enter the name you want the output file to be: ')+'.xlsx'
path=""
inp_name=""
count=0

for i in range(len(source_inp)-1,-1,-1):
    if source_inp[i]!='\\' and count==0:
        inp_name=source_inp[i]+inp_name
    else:
        count=1
    if count==1:
        if source_inp[i]=='\\':
            path='/'+path
        else:
            path=source_inp[i]+path

wb = Workbook()
wb.save(filename=destination)
wb.save(filename='rejected.xlsx')

''''''''''''''''''''''''''''''''''''''''''


# loading the source xl
actual = load_workbook(source_inp[1:len(source_inp)-1])
take_from = actual.active


# loading the output xl
to = load_workbook(destination)
write_to = to.active

# loading the output xl
rest = load_workbook('rejected.xlsx')
rejected_write = rest.active

# variable empty for counting empty cells and count for taking inpur row-wise
# dic_store for future sorting by district name
# writecount for writing the final result
empty = 0
count = 1
dic_store={}
writecount = 1
rejected = 1

# the loop will end if it finds 100 consecutive rows containing nothing
while empty < 100:
# let the fun begin :-)
    # taking inputs
    serial = take_from[f'a{count}'].value
    membership = take_from[f'b{count}'].value
    name = take_from[f'c{count}'].value

    # removing unwanted items from the name
    try:
        while True:

            if "(" in name:

                start=name.index('(')
                end=name.index(')')
                new=''
                for i in range(len(name)):
                    if i <start or  i>end:
                        new+=name[i]
                name=new
            elif 'LIFE MEMBER' in name:
                name=name.replace('LIFE MEMBER','')
            else:
                break  
    except:
        pass
    
    # adding prefix and suffix to the name
    try:
        name=name.strip()
        name="Ld. "+ name +' Sir, '
    except:
        pass
    
    info = take_from[f'd{count}'].value
    dist = take_from[f'f{count}'].value
    extra = str(take_from[f'e{count}'].value).strip()
    ex=extra

    # sorting the extra numbers according to different parameters
    if ', '  in extra:
        extra=extra.split(', ')
    elif ',' in extra:
        extra=extra.split(',')
    elif '/' in extra:
        extra=extra.split('/')
    elif '-' in extra and len(extra)==12:
        extra=[extra[0:5]+extra[6:12],'r']
    elif '-' in extra and len(extra)==23:
        extra=[extra[0:11],extra[12:23],'s']
    elif len(extra)==22:
        extra=[extra[0:11],extra[11:22],'s']
    else:
        extra=extra.split()

    count += 1
    # end of taking inputs------ count+=1 for taking input from next rows

    # condition inside "if" is for an empty row 
    if name == None and info == None:
        empty += 1
    else:
        # empty 0 for resseting the value and numlis is tha ultimate storage for numbers
        empty = 0
        numlis = []
        if info != None:
            try:
                # traversing the 'info' if we encounter '0' a loop will start for taking the next 11 strings
                for i in range(len(info)):
                    if info[i] == '0':
                        num = ''
                        try:
                            # the loop to find the number-------usng try except method to avoid out of range error
                            for j in range(i, i + 11):
                                num += info[j]
                            try:
                                # if it is convertable to intiger then ok
                                # using try except for avoiding hybrid strings or incorrect strings
                                a = int(num)
                                numlis.append(num.strip())
                            except:
                                pass
                        except:
                            pass
            except:
                pass
        # fixing the extra list.... cause inputs are fucked and sepereted by various sorts of shits
        for i in range(len(extra)):
            if len(extra[i]) == 10 and extra[i][0] == '1':
                extra[i] = '0' + extra[i]
            if len(extra[i]) == 12 and extra[i][-1] == '.':
                extra[i] = extra[i][0:11]
            if len(extra[i]) == 12 and extra[i][-1] == ',':
                extra[i] = extra[i][0:11]
            if len(extra[i]) == 12 and extra[i][0] == '`':
                extra[i] = extra[i][1:12]

        # finally adding the valid inputs to the numlis
        for i in extra:
            if len(i) == 11 and i not in numlis:
                numlis.append(i)

        # replacing any invalid numbers of citycell numbers with '0'
        for i in numlis:
            if (len(i) != 11) or ('011' == i[0:3]):
                numlis[numlis.index(i)] = 0

        # creating new numlis without '0's  --- in other words without any invalid numbers
        numlis = [i for i in numlis if i != 0]

        # just taking the last valid number
        if numlis!=[]:
            numlis=[numlis[-1]]

        # finally writing the valid numbers in the desired file
        if numlis != []:
            lala = 0 # to write the name of the district just once

            for i in numlis:
                if lala == 0:
                    write_to[f'd{writecount}'] = info
                    lala = 1
                write_to[f'a{writecount}'] = serial
                write_to[f'b{writecount}'] = membership
                write_to[f'c{writecount}'] = name
                write_to[f'e{writecount}'] = i
                write_to[f'f{writecount}'] = dist

                writecount += 1
        else:
            rejected_write[f'a{rejected}'] = serial
            rejected_write[f'b{rejected}'] = membership
            rejected_write[f'c{rejected}'] = name
            rejected_write[f'd{rejected}'] = info
            rejected_write[f'e{rejected}'] = ex
            rejected_write[f'f{rejected}'] = dist
            rejected+=1

to.save(destination)
rest.save('rejected.xlsx')
